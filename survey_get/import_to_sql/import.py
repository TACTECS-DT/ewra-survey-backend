import os
from openpyxl import load_workbook




def excel_directory_to_sql(directory_path, output_path):
    for filename in os.listdir(directory_path):
        if filename.endswith(".xlsx"):
            excel_path = os.path.join(directory_path, filename)
            table_name = 'survey_app_'+ os.path.splitext(filename)[0].lower()  

            wb = load_workbook(excel_path)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            headers = rows[0]
            insert_statements = []

            for row in rows[1:]:  # Skip header
                values = []
                for value in row:
                    if value is None:
                        values.append("NULL")
                    elif isinstance(value, (int, float)):
                        values.append(str(value))
                    else:
                        # Escape single quotes
                        safe_value = str(value).replace("'", "''")
                        values.append(f"'{safe_value}'")
                columns = ', '.join(headers)
                values_str = ', '.join(values)
                insert = f"INSERT INTO {table_name} ({columns}) VALUES ({values_str});"
                insert_statements.append(insert)

            # Write SQL statements to .sql file
            sql_file_path = os.path.join(output_path, f"{table_name}.sql")
            with open(sql_file_path, "w", encoding="utf-8") as f:
                for stmt in insert_statements:
                    f.write(stmt + "\n")

            print(f"Generated: {sql_file_path}")

excel_directory_to_sql("D:\work\django_production_apps\survey\import_to_sql\input", "D:\work\django_production_apps\survey\import_to_sql\output")