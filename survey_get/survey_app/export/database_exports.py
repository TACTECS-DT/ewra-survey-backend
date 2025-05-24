
from django.db import connection
from openpyxl import Workbook
from ..all_models import main as base_models 
from ..all_models import survey as survey_models
import json
from django.http import HttpResponse, Http404 ,HttpResponseBadRequest
from zoneinfo import ZoneInfo  
from datetime import datetime
from django.utils import timezone
from django.contrib.auth.decorators import permission_required , login_required
from django.db.models import Count
from..custom_user import CustomUser
from django.core.exceptions import ValidationError
from ..utils import is_role_admin  
from django.apps import apps
import zipfile
import io
from django.db import models

# List of allowed models
ALLOWED_MODELS = [
    'Governorate', 'Center', 'ServiceProvider', 'Affiliate', 'Branch',
    'MainActivity', 'ActivityType', 'ActivityItem', 'MainSurveySection',
    'SubSurveySection', 'QuestionBank', 'Survey', 'SurveyQuestion',
    # 'SurveyQuestionsUpdateHistory'
]





@login_required
@is_role_admin
def export_many2many_to_excel(request):
    table_name = "survey_app_questionbank_sub_survey_sections"
    columns = ["*"] 
    return _export_many2many_table_to_excel(table_name, columns)


def _get_columns_from_table(table_name):
    # Fetch the columns dynamically from the database
    with connection.cursor() as cursor:
        cursor.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_name}'")
        columns = cursor.fetchall()

    # Extract column names from the result
    column_names = [column[0] for column in columns]
    return column_names


def _export_many2many_table_to_excel(table_name, columns):
    # Dynamically fetch column names from the database if columns aren't provided or are incomplete
    if columns == ['*']:  # If columns list contains a single wildcard ('*'), fetch all columns
        columns = _get_columns_from_table(table_name)

    # Execute raw SQL to fetch data from the table
    column_list_str = ', '.join(columns)  # Prepare the column list string for the SQL query
    with connection.cursor() as cursor:
        cursor.execute(f"SELECT {column_list_str} FROM {table_name}")
        records = cursor.fetchall()

    # Check if records are empty
    if not records:
        return HttpResponse("No data found in the table.", status=404)

    # Create a new workbook and add the data
    wb = Workbook()
    ws = wb.active
    ws.title = table_name
    ws.append(columns)  # Write the header row with the dynamic columns
    file_name = 'm2m'+str(table_name)
    # Write data rows
    for row in records:
        ws.append(row)

    # Return the Excel file as a response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
    wb.save(response)
    return response


@login_required
@is_role_admin
def export_model_to_excel(request):
    model_param = request.GET.get('model')

    if not model_param:
        return HttpResponseBadRequest("Model name is required.")

    if model_param == "all":
        return _export_all_models_to_zip()
    elif model_param not in ALLOWED_MODELS:
        return Http404("Invalid model")

    return _export_single_model_to_excel(model_param)



@login_required
@is_role_admin
def export_model_to_excel(request):
    model_param = request.GET.get('model')

    if not model_param:
        return HttpResponseBadRequest("Model name is required.")

    if model_param == "all":
        return _export_all_models_to_zip()
    elif model_param not in ALLOWED_MODELS:
        return Http404("Invalid model")

    return _export_single_model_to_excel(model_param)


def _export_single_model_to_excel(model_name):
    model = apps.get_model('survey_app', model_name)  # Replace with your actual app label
    records = model.objects.all()
    # field_names = [field.name for field in model._meta.fields]
    field_names = [
        f"{field.name}_id" if isinstance(field, models.ForeignKey) else field.name
        for field in model._meta.fields
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = model_name
    ws.append(field_names)

    for rec in records:
        row = [str(getattr(rec, field)) if getattr(rec, field) is not None else '' for field in field_names]
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={model_name}.xlsx'
    wb.save(response)
    return response


def _export_all_models_to_zip():
    
    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, 'w') as zip_file:
        for model_name in ALLOWED_MODELS:
            model = apps.get_model('survey_app', model_name)  # Replace with your app label
            records = model.objects.all()
            # field_names = [field.name for field in model._meta.fields]
            field_names = [
                    f"{field.name}_id" if isinstance(field, models.ForeignKey) else field.name
                    for field in model._meta.fields
                ]
            wb = Workbook()
            ws = wb.active
            ws.title = model_name
            ws.append(field_names)

            for rec in records:
                row = [str(getattr(rec, field)) if getattr(rec, field) is not None else '' for field in field_names]
                ws.append(row)

            # Save each workbook to memory and add it to the zip
            excel_io = io.BytesIO()
            wb.save(excel_io)
            excel_io.seek(0)
            zip_file.writestr(f"{model_name}.xlsx", excel_io.read())

    memory_file.seek(0)
    response = HttpResponse(memory_file, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=All_Models_Export.zip'
    return response




